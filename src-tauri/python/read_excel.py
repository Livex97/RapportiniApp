#!/usr/bin/env python3
import sys
import json
import openpyxl
import os

def is_valid_header(value):
    """Verifica se l'intestazione è valida (non vuota e non solo spazi)."""
    if value is None:
        return False
    str_val = str(value).strip()
    return len(str_val) > 0

def find_sheet(wb, type_hint):
    """Cerca il foglio corretto basandosi sul tipo (pandetta o sterlink)."""
    if type_hint == 'pandetta':
        # Cerca fogli che contengono "PANDETTA" nel nome
        for sheet_name in wb.sheetnames:
            if "PANDETTA" in sheet_name.upper():
                return wb[sheet_name]
    elif type_hint == 'sterlink':
        # Sterlink di solito ha un solo foglio o il primo
        return wb.active
    
    # Fallback al primo foglio
    return wb.active

def main():
    if len(sys.argv) < 3:
        print("Usage: read_excel.py <file_path> <type_hint>", file=sys.stderr)
        sys.exit(1)

    file_path = sys.argv[1]
    type_hint = sys.argv[2]

    if not os.path.exists(file_path):
        print(f"Error: File not found at {file_path}", file=sys.stderr)
        sys.exit(1)

    try:
        # Carica solo dati (data_only=True per risolvere le formule)
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = find_sheet(wb, type_hint)
        
        # Estrai intestazioni (riga 1) - salta colonne senza intestazione
        header_cells = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        valid_indices = [i for i, h in enumerate(header_cells) if is_valid_header(h)]
        headers = [str(header_cells[i]).strip() for i in valid_indices]
        
        rows = []
        # Determina il numero massimo di colonne con intestazione valida
        max_col = max(valid_indices) + 1 if valid_indices else 0
        
        # Estrai dati (dalla riga 2 in poi)
        for row_cells in ws.iter_rows(min_row=2, values_only=True):
            # Salta righe completamente vuote
            if all(c is None for c in row_cells[:max_col]):
                continue
            
            row_dict = {}
            for i in valid_indices:
                if i < len(row_cells):
                    val = row_cells[i]
                    # Formatta valori (es. date in stringhe ISO)
                    from datetime import datetime
                    if isinstance(val, datetime):
                        val = val.strftime('%Y-%m-%d')
                    row_dict[headers[valid_indices.index(i)]] = val
            rows.append(row_dict)

        # Output strutturato: colonne ordinate e righe
        output_data = {
            "columns": headers,
            "rows": rows
        }
        print(json.dumps(output_data))
        
    except Exception as e:
        print(f"Error reading excel: {e}", file=sys.stderr)
        sys.exit(1)

if __name__ == '__main__':
    main()
